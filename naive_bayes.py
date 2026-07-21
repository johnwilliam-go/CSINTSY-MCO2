import os
from sklearn.naive_bayes import CategoricalNB
from sklearn.preprocessing import OrdinalEncoder, LabelEncoder
from openpyxl import load_workbook
import pandas as pd
import numpy as np


#-- Reading the xlsx file --
class Readxlsx:
    def __init__(self, filename: str):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        self.filename = os.path.join(script_dir, filename)

    def reading(self):
        if not os.path.exists(self.filename):
            raise FileNotFoundError(f"Could not find file {self.filename}")

        wb = load_workbook(self.filename)

        aws = wb.active

        raw_data = [row[0] for row in aws.iter_rows(min_row=345, max_row=618, min_col=4, max_col=4, values_only=True)
                    ]

        # Convert into NumPy Array
        raw_data_array = np.array(raw_data)

        return raw_data_array

#-- Convert into feature matrix --
class FeatureMatrix:
    def __init__(self, filename: str):
        self.filename = filename
        self.read_xlsx = Readxlsx(filename)

    def extracting_features(self, word: str) -> dict:
        clean_word = word.lower().strip()

        #-- Feature Columns --
        fil_prefixes = ('pinaka', 'pakikipag', 'pagkaka',
                        'pakiki', 'pagka', 'maka', 'paki',
                        'mag', 'nag', 'pag', 'ipa', 'um', 'in')


        fil_suffixes = ('an', 'in', 'han', 'hin', 'ng')

        fil_prepositions = ('nasa', 'sa')

        eng_prepositions = ('from', 'among', 'towards', 'across',
                            'through', 'between', 'of', 'with', 'by',
                            'about', 'for')

        eng_prefix = ('anti', 'de', 'dis', 'en', 'em',
                      'fore', 'inter', 'mid', 'mis', 'non',
                      'over', 'pre', 're', 'semi', 'sub', 'super',
                      'trans', 'un', 'under')

        eng_suffixes = ('able', 'ible', 'al', 'ial', 'ed', 'en', 'er',
                        'est', 'ful', 'ic', 'ing', 'ion', 'tion',
                        'ation', 'ition', 'ity', 'ty', 'ive', 'ative', 'itive',
                        'less', 'ly', 'ment', 'ness', 'ous', 'eous', 'ious',
                        's', 'es', 'y')

        ng_or_nang = ('ng', 'nang')

        punctuation_marks = ('.', ',', '!', '?', '-')

        has_fil_prefix = 1 if clean_word.startswith(fil_prefixes) else 0

        has_fil_suffix = 1 if clean_word.endswith(fil_suffixes) else 0

        has_fil_prepositions = 1 if clean_word in fil_prepositions else 0

        has_eng_prepositions = 1 if clean_word in eng_prepositions else 0

        has_eng_prefix = 1 if clean_word.startswith(eng_prefix) else 0

        has_eng_suffix = 1 if clean_word.startswith(eng_suffixes) else 0

        has_ng_or_nang = 1 if clean_word in ng_or_nang else 0

        code_switching = 1 if clean_word.startswith(fil_prefixes) and (clean_word.endswith(eng_suffixes) or clean_word.startswith(eng_prefix)) or ("-" in clean_word and clean_word.startswith(fil_prefixes)) else 0

        has_punctuation_marks = 1 if clean_word in punctuation_marks else 0

        return {'Fil Prefix': has_fil_prefix,
                'Fil Suffix': has_fil_suffix,
                'Fil Prepositions': has_fil_prepositions,
                'Eng Prefix': has_eng_prefix,
                'Eng Suffix': has_eng_suffix,
                'Eng Prepositions': has_eng_prepositions,
                'Ng or Nang': has_ng_or_nang,
                'Code Switching': code_switching,
                'Punctuation Marks': has_punctuation_marks}

    def converting_to_feature_matrix(self):
        raw_words = self.read_xlsx.reading()

        feature_words = [] # Dataframe column
        index_words = [] # Dataframe row

        for word in raw_words:
            features = self.extracting_features(word)
            feature_words.append(features)
            index_words.append(word)

        df = pd.DataFrame(feature_words, index=index_words)

        return df


class NaivePrediction:
    def __init__(self):
        self.rdata_label_encoder = LabelEncoder()
        self.rdata_ordinal_encoder = OrdinalEncoder()
        self.model_categoricalnb = CategoricalNB()

        # -- Encoded target labels for model training --
        target_labels = ['ENG', 'FIL', 'CS', 'OTH']
        fitted_tlabels = self.rdata_label_encoder.fit(target_labels)

    def CategoricalToInteger(self, dataframe):
        encoded_dataframe = self.rdata_ordinal_encoder.fit(dataframe)

        return encoded_dataframe

    def TrainModel(self, encoded_dataframe):
        fitted_data = self.model_categoricalnb.fit(encoded_dataframe, self.fitted_tlabels)

        return fitted_data

    def PredictedValueOfToken(self, token: str):
        pass




#-- Driver for testing --
data = input("Input file name: ").strip()

fm = FeatureMatrix(data)

data_frame = fm.converting_to_feature_matrix()

nb = NaivePrediction()

X_fit = nb.CategoricalToInteger(data_frame)

#-- Don't mind the block of code below
#rx = Readxlsx(data)

#data_array = rx.reading()

#df = pd.DataFrame(data_array, columns=["word"])

#print(df)